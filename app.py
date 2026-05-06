import os
import sys
import json
import subprocess
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
 
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
 
TASK_NAME = "MRDR_Report_Updater_Daily"
SETTINGS_FILE = "settings.json"
 
COL_ARTICLE = "Артикул (MRDR) продукта"
COL_CREATE_DATE = "Дата создания"
COL_CREATE_DATE_RAW = "Дата создания (raw)"
COL_PROGRESS = "Значения Прогресса"
 
EXTRA_COLUMNS_PATTERNS = {
    "Perfect name": ["Perfect name"],
    "Division description": ["Division description"],
    "Категория продукта": ["Категория продукта"],
    "Бренд": ["Бренд"],
    "Название бренда": ["Название бренда"],
}
 
REPORT_SHEET_NAME = "Лист1"
REPORT_HEADER_ROW = 2
REPORT_DATA_START_ROW = 3
 
ARTICLE_COL_IDX = 1
CREATE_DATE_COL_IDX = 2
 
def app_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))
 
def settings_path() -> str:
    return os.path.join(app_dir(), SETTINGS_FILE)
 
def load_settings() -> dict:
    path = settings_path()
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}
 
def save_settings(data: dict):
    with open(settings_path(), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
 
def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\xa0", " ").strip().lower()
 
def normalize_article(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    if text.lower() == "nan":
        return ""
    return text
 
def get_export_file_modified_datetime(export_path: str) -> datetime:
    ts = os.path.getmtime(export_path)
    return datetime.fromtimestamp(ts)
 
def center_alignment(wrap_text: bool = False) -> Alignment:
    return Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=wrap_text
    )
 
def get_min_allowed_date() -> pd.Timestamp:
    current_year = datetime.now().year
    previous_year = current_year - 1
    return pd.Timestamp(year=previous_year, month=1, day=1)
 
def parse_create_date(value):
    if value is None:
        return pd.NaT
 
    if isinstance(value, (pd.Timestamp, datetime)):
        return pd.Timestamp(value)
 
    if pd.isna(value):
        return pd.NaT
 
    text = str(value).strip().replace("\xa0", " ")
    if not text:
        return pd.NaT
 
    for fmt in (
        "%d.%m.%Y",
        "%Y-%m-%d",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S"
    ):
        try:
            return pd.to_datetime(text, format=fmt, errors="raise")
        except Exception:
            pass
 
    try:
        return pd.to_datetime(text, errors="coerce", dayfirst=True)
    except Exception:
        return pd.NaT
 
def is_date_in_allowed_range(value) -> bool:
    dt = parse_create_date(value)
    if pd.isna(dt):
        return False
    return dt >= get_min_allowed_date()
 
def find_column_by_patterns(columns, patterns):
    normalized_columns = {col: normalize_text(col) for col in columns}
 
    for pattern in patterns:
        pattern_norm = normalize_text(pattern)
        for original_col, normalized_col in normalized_columns.items():
            if normalized_col == pattern_norm:
                return original_col
 
    return None
 
def read_export_file(export_path: str):
    df = pd.read_excel(
        export_path,
        sheet_name="products",
        header=0,
        dtype=object
    )
 
    normalized_columns = {col: normalize_text(col) for col in df.columns}
 
    article_col = None
    create_date_col = None
    progress_col = None
 
    for original, normalized in normalized_columns.items():
        if article_col is None and "артикул" in normalized and "mrdr" in normalized:
            article_col = original
        elif create_date_col is None and "дата" in normalized and "создания" in normalized:
            create_date_col = original
        elif progress_col is None and "прогресс" in normalized:
            progress_col = original
 
    extra_columns_found = {}
    missing_extra_columns = []
 
    for target_name, patterns in EXTRA_COLUMNS_PATTERNS.items():
        found_col = find_column_by_patterns(df.columns, patterns)
        if found_col is not None:
            extra_columns_found[target_name] = found_col
        else:
            missing_extra_columns.append(target_name)
 
    missing = []
    if not article_col:
        missing.append(COL_ARTICLE)
    if not create_date_col:
        missing.append(COL_CREATE_DATE)
    if not progress_col:
        missing.append(COL_PROGRESS)
 
    missing.extend(missing_extra_columns)
 
    if missing:
        raise ValueError(
            "Не найдены нужные колонки в выгрузке:\n- " + "\n- ".join(missing)
        )
 
    selected_cols = [article_col, create_date_col] + list(extra_columns_found.values()) + [progress_col]
    result = df[selected_cols].copy()
 
    renamed_columns = {
        article_col: COL_ARTICLE,
        create_date_col: COL_CREATE_DATE,
        progress_col: COL_PROGRESS
    }
 
    for target_name, source_col in extra_columns_found.items():
        renamed_columns[source_col] = target_name
 
    result.rename(columns=renamed_columns, inplace=True)
 
    result[COL_CREATE_DATE_RAW] = result[COL_CREATE_DATE]
    result[COL_ARTICLE] = result[COL_ARTICLE].apply(normalize_article)
    result[COL_CREATE_DATE] = result[COL_CREATE_DATE].apply(parse_create_date)
 
    min_allowed_date = get_min_allowed_date()
 
    result = result[
        (result[COL_ARTICLE] != "") &
        (result[COL_CREATE_DATE].notna()) &
        (result[COL_CREATE_DATE] >= min_allowed_date)
    ].copy()
 
    result = result.dropna(subset=[COL_ARTICLE])
    result = result.drop_duplicates(subset=[COL_ARTICLE], keep="last").copy()
 
    result.reset_index(drop=True, inplace=True)
    return result, list(extra_columns_found.keys())
 
def ensure_report_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    return ws
 
def initialize_report_headers(ws):
    ws.cell(row=REPORT_HEADER_ROW, column=ARTICLE_COL_IDX, value=COL_ARTICLE)
    ws.cell(row=REPORT_HEADER_ROW, column=CREATE_DATE_COL_IDX, value=COL_CREATE_DATE)
 
    ws.cell(row=REPORT_HEADER_ROW, column=ARTICLE_COL_IDX).alignment = center_alignment(wrap_text=True)
    ws.cell(row=REPORT_HEADER_ROW, column=CREATE_DATE_COL_IDX).alignment = center_alignment(wrap_text=True)
 
def get_report_headers_map(ws) -> dict:
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=REPORT_HEADER_ROW, column=col_idx).value
        if val is None:
            continue
        header_text = str(val).strip()
        if header_text:
            headers[header_text] = col_idx
    return headers
 
def shift_progress_columns_if_needed(ws, extra_headers: list):
    progress_columns = []
 
    for col_idx in range(3, ws.max_column + 1):
        header_value = ws.cell(row=REPORT_HEADER_ROW, column=col_idx).value
        header_text = "" if header_value is None else str(header_value).strip()
 
        if header_text == COL_PROGRESS:
            progress_columns.append(col_idx)
 
    if not progress_columns:
        return
 
    first_progress_col = min(progress_columns)
    required_first_progress_col = 3 + len(extra_headers)
 
    if first_progress_col >= required_first_progress_col:
        return
 
    shift_by = required_first_progress_col - first_progress_col
    ws.insert_cols(first_progress_col, amount=shift_by)
 
def ensure_extra_columns(ws, extra_headers: list) -> dict:
    headers_map = get_report_headers_map(ws)
    extra_col_map = {}
 
    next_col_idx = 3
    while ws.cell(row=REPORT_HEADER_ROW, column=next_col_idx).value not in (None, ""):
        next_col_idx += 1
 
    for header in extra_headers:
        if header in headers_map:
            extra_col_map[header] = headers_map[header]
        else:
            ws.cell(row=REPORT_HEADER_ROW, column=next_col_idx, value=header)
            ws.cell(row=REPORT_HEADER_ROW, column=next_col_idx).alignment = center_alignment(wrap_text=True)
            extra_col_map[header] = next_col_idx
            next_col_idx += 1
 
    return extra_col_map
 
def find_or_create_date_column(ws, export_modified_dt: datetime, first_progress_col_idx: int) -> int:
    target_dt_str = export_modified_dt.strftime("%d.%m.%Y\n%H:%M")
 
    col_idx = first_progress_col_idx
    while True:
        header_value = ws.cell(row=REPORT_HEADER_ROW, column=col_idx).value
        top_value = ws.cell(row=1, column=col_idx).value
 
        header_str = "" if header_value is None else str(header_value).strip()
        top_str = "" if top_value is None else str(top_value).strip()
 
        if header_str == "" and top_str == "":
            break
 
        if header_str == COL_PROGRESS and top_str == target_dt_str:
            return col_idx
 
        col_idx += 1
 
    new_col_idx = col_idx
 
    top_cell = ws.cell(row=1, column=new_col_idx, value=target_dt_str)
    top_cell.alignment = center_alignment(wrap_text=True)
 
    header_cell = ws.cell(row=REPORT_HEADER_ROW, column=new_col_idx, value=COL_PROGRESS)
    header_cell.alignment = center_alignment(wrap_text=True)
 
    return new_col_idx
 
def get_existing_articles_map(ws) -> dict:
    article_map = {}
 
    for row_idx in range(REPORT_DATA_START_ROW, ws.max_row + 1):
        article_value = ws.cell(row=row_idx, column=ARTICLE_COL_IDX).value
        article = normalize_article(article_value)
        if article:
            article_map[article] = row_idx
 
    return article_map
 
def find_next_empty_row(ws) -> int:
    row_idx = REPORT_DATA_START_ROW
    while True:
        if not normalize_article(ws.cell(row=row_idx, column=ARTICLE_COL_IDX).value):
            return row_idx
        row_idx += 1
 
def apply_alignment_to_row(ws, row_idx: int):
    for col_idx in range(1, ws.max_column + 1):
        wrap = row_idx in (1, REPORT_HEADER_ROW)
        ws.cell(row=row_idx, column=col_idx).alignment = center_alignment(wrap_text=wrap)
 
def apply_alignment_to_all(ws):
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            wrap = row_idx in (1, REPORT_HEADER_ROW)
            ws.cell(row=row_idx, column=col_idx).alignment = center_alignment(wrap_text=wrap)
 
def auto_fit_columns(ws, extra_col_indexes=None, first_progress_col_idx=None):
    extra_col_indexes = set(extra_col_indexes or [])
 
    for col_idx in range(1, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
 
        if col_idx == ARTICLE_COL_IDX:
            ws.column_dimensions[col_letter].width = 18
            continue
 
        if col_idx == CREATE_DATE_COL_IDX:
            ws.column_dimensions[col_letter].width = 16
            continue
 
        if col_idx in extra_col_indexes:
            ws.column_dimensions[col_letter].width = 22
            continue
 
        if first_progress_col_idx is not None and col_idx >= first_progress_col_idx:
            header_value = ws.cell(row=REPORT_HEADER_ROW, column=col_idx).value
            if str(header_value).strip() == COL_PROGRESS:
                ws.column_dimensions[col_letter].width = 12
                continue
 
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            text = "" if val is None else str(val)
            if len(text) > max_len:
                max_len = len(text)
 
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)
 
def remove_outdated_rows(ws) -> int:
    rows_to_delete = []
 
    for row_idx in range(REPORT_DATA_START_ROW, ws.max_row + 1):
        article = normalize_article(ws.cell(row=row_idx, column=ARTICLE_COL_IDX).value)
        create_date_value = ws.cell(row=row_idx, column=CREATE_DATE_COL_IDX).value
 
        if not article:
            continue
 
        if not is_date_in_allowed_range(create_date_value):
            rows_to_delete.append(row_idx)
 
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx, 1)
 
    return len(rows_to_delete)
 
def write_text_date(cell, raw_value):
    if raw_value is None:
        return
 
    value = str(raw_value).strip()
    if not value:
        return
 
    cell.number_format = "@"
    cell.value = value
 
def write_report(report_path: str, data: pd.DataFrame, export_modified_dt: datetime, extra_headers: list):
    if os.path.exists(report_path):
        wb = load_workbook(report_path)
    else:
        wb = Workbook()
 
    ws = ensure_report_sheet(wb, REPORT_SHEET_NAME)
 
    initialize_report_headers(ws)
 
    removed_count = remove_outdated_rows(ws)
 
    shift_progress_columns_if_needed(ws, extra_headers)
    extra_col_map = ensure_extra_columns(ws, extra_headers)
 
    first_progress_col_idx = 3 + len(extra_headers)
    progress_col_idx = find_or_create_date_column(ws, export_modified_dt, first_progress_col_idx)
 
    ws.cell(row=1, column=progress_col_idx).alignment = center_alignment(wrap_text=True)
    ws.cell(row=REPORT_HEADER_ROW, column=progress_col_idx).alignment = center_alignment(wrap_text=True)
 
    article_map = get_existing_articles_map(ws)
 
    added_count = 0
    updated_count = 0
 
    for _, row in data.iterrows():
        article = normalize_article(row[COL_ARTICLE])
        create_date_raw = row.get(COL_CREATE_DATE_RAW)
        progress = row[COL_PROGRESS]
 
        if not article:
            continue
 
        if article in article_map:
            row_idx = article_map[article]
            updated_count += 1
        else:
            row_idx = find_next_empty_row(ws)
            ws.cell(row=row_idx, column=ARTICLE_COL_IDX, value=article)
            write_text_date(ws.cell(row=row_idx, column=CREATE_DATE_COL_IDX), create_date_raw)
            article_map[article] = row_idx
            added_count += 1
 
        current_date_cell_value = ws.cell(row=row_idx, column=CREATE_DATE_COL_IDX).value
        if current_date_cell_value is None or str(current_date_cell_value).strip() == "":
            write_text_date(ws.cell(row=row_idx, column=CREATE_DATE_COL_IDX), create_date_raw)
 
        for header in extra_headers:
            value = row.get(header, None)
            ws.cell(row=row_idx, column=extra_col_map[header], value=value)
 
        ws.cell(row=row_idx, column=progress_col_idx, value=progress)
 
        apply_alignment_to_row(ws, row_idx)
 
    apply_alignment_to_all(ws)
 
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[REPORT_HEADER_ROW].height = 34
 
    for row_idx in range(REPORT_DATA_START_ROW, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22
 
    auto_fit_columns(
        ws,
        extra_col_indexes=list(extra_col_map.values()),
        first_progress_col_idx=first_progress_col_idx
    )
 
    wb.save(report_path)
 
    return added_count, updated_count, removed_count, progress_col_idx
 
def update_report(export_path: str, report_path: str):
    if not export_path or not os.path.exists(export_path):
        raise FileNotFoundError("Файл выгрузки не найден.")
    if not report_path:
        raise FileNotFoundError("Не указан путь к отчету.")
    if not os.path.exists(report_path):
        raise FileNotFoundError("Файл отчета не найден. Нужно выбрать уже существующий отчет.")
 
    export_modified_dt = get_export_file_modified_datetime(export_path)
    data, extra_headers = read_export_file(export_path)
 
    if data.empty:
        raise ValueError(
            "После чтения выгрузки не осталось ни одной строки.\n"
            "Проверь:\n"
            "- есть ли артикулы,\n"
            "- правильно ли читается 'Дата создания',\n"
            "- не отсекает ли всё фильтр по текущему и предыдущему году."
        )
 
    added_count, updated_count, removed_count, progress_col_idx = write_report(
        report_path, data, export_modified_dt, extra_headers
    )
 
    return {
        "rows_in_export": len(data),
        "added_count": added_count,
        "updated_count": updated_count,
        "removed_count": removed_count,
        "export_modified_dt": export_modified_dt,
        "progress_col_idx": progress_col_idx,
        "extra_cols_count": len(extra_headers)
    }
 
def create_runner_bat():
    bat_path = os.path.join(app_dir(), "run_update.bat")
 
    if getattr(sys, "frozen", False):
        cmd = f'"{sys.executable}" --autoupdate'
    else:
        script_path = os.path.abspath(__file__)
        cmd = f'"{sys.executable}" "{script_path}" --autoupdate'
 
    content = f"""@echo off
cd /d "{app_dir()}"
{cmd}
"""
    with open(bat_path, "w", encoding="utf-8") as f:
        f.write(content)
 
    return bat_path
 
def install_task(run_time: str):
    try:
        datetime.strptime(run_time, "%H:%M")
    except ValueError:
        raise ValueError("Время должно быть в формате ЧЧ:ММ, например 09:30")
 
    bat_path = create_runner_bat()
 
    cmd = [
        "schtasks",
        "/Create",
        "/F",
        "/SC", "DAILY",
        "/TN", TASK_NAME,
        "/TR", bat_path,
        "/ST", run_time
    ]
 
    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        shell=False
    )
 
    if result.returncode != 0:
        raise RuntimeError(
            result.stderr.strip() or
            result.stdout.strip() or
            "Не удалось создать задачу в планировщике."
        )
 
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Обновление прогресса")
        self.geometry("820x300")
        self.resizable(False, False)
 
        self.export_path_var = tk.StringVar()
        self.report_path_var = tk.StringVar()
        self.time_var = tk.StringVar(value="09:00")
        self.status_var = tk.StringVar(value="Примечание: данные тянутся за текущий и прошлый год")
 
        self.build_ui()
        self.load_form_settings()
 
    def build_ui(self):
        pad = {"padx": 10, "pady": 8}
 
        main = ttk.Frame(self)
        main.pack(fill="both", expand=True, padx=12, pady=12)
 
        ttk.Label(main, text="Путь к выгрузке:").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(main, textvariable=self.export_path_var, width=78).grid(row=0, column=1, sticky="we", **pad)
        ttk.Button(main, text="Обзор", command=self.browse_export).grid(row=0, column=2, **pad)
 
        ttk.Label(main, text="Путь к отчету:").grid(row=1, column=0, sticky="w", **pad)
        ttk.Entry(main, textvariable=self.report_path_var, width=78).grid(row=1, column=1, sticky="we", **pad)
        ttk.Button(main, text="Обзор", command=self.browse_report).grid(row=1, column=2, **pad)
 
        ttk.Label(main, text="Время ежедневного запуска (ЧЧ:ММ):").grid(row=2, column=0, sticky="w", **pad)
        ttk.Entry(main, textvariable=self.time_var, width=20).grid(row=2, column=1, sticky="w", **pad)
 
        btns = ttk.Frame(main)
        btns.grid(row=3, column=0, columnspan=3, pady=20)
 
        ttk.Button(btns, text="Сохранить настройки", command=self.on_save_settings).pack(side="left", padx=8)
        ttk.Button(btns, text="Установить автозапуск", command=self.on_install_autorun).pack(side="left", padx=8)
        ttk.Button(btns, text="Обновить отчет", command=self.on_update_report).pack(side="left", padx=8)
 
        ttk.Label(main, textvariable=self.status_var, foreground="blue").grid(
            row=4, column=0, columnspan=3, sticky="w", padx=10, pady=10
        )
 
        main.columnconfigure(1, weight=1)
 
    def browse_export(self):
        path = filedialog.askopenfilename(
            title="Выберите файл выгрузки",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self.export_path_var.set(path)
 
    def browse_report(self):
        path = filedialog.askopenfilename(
            title="Выберите существующий файл отчета",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self.report_path_var.set(path)
 
    def load_form_settings(self):
        s = load_settings()
        self.export_path_var.set(s.get("export_path", ""))
        self.report_path_var.set(s.get("report_path", ""))
        self.time_var.set(s.get("run_time", "09:00"))
 
    def collect_settings(self) -> dict:
        return {
            "export_path": self.export_path_var.get().strip(),
            "report_path": self.report_path_var.get().strip(),
            "run_time": self.time_var.get().strip()
        }
 
    def on_save_settings(self):
        try:
            data = self.collect_settings()
 
            if not data["export_path"]:
                raise ValueError("Укажите путь к выгрузке.")
            if not os.path.exists(data["export_path"]):
                raise ValueError("Файл выгрузки не найден.")
            if not data["report_path"]:
                raise ValueError("Укажите путь к отчету.")
            if not os.path.exists(data["report_path"]):
                raise ValueError("Файл отчета не найден. Нужно выбрать уже существующий отчет.")
 
            datetime.strptime(data["run_time"], "%H:%M")
 
            save_settings(data)
            self.status_var.set("Настройки сохранены")
            messagebox.showinfo("Успешно", "Настройки сохранены.")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
 
    def on_install_autorun(self):
        try:
            data = self.collect_settings()
 
            if not data["export_path"]:
                raise ValueError("Сначала укажи путь к выгрузке.")
            if not os.path.exists(data["export_path"]):
                raise ValueError("Файл выгрузки не найден.")
            if not data["report_path"]:
                raise ValueError("Сначала укажи путь к отчету.")
            if not os.path.exists(data["report_path"]):
                raise ValueError("Файл отчета не найден. Нужно выбрать уже существующий отчет.")
 
            save_settings(data)
            install_task(data["run_time"])
 
            self.status_var.set(f"Автозапуск установлен на {data['run_time']}")
            messagebox.showinfo("Успешно", f"Автозапуск установлен на {data['run_time']}.")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
 
    def on_update_report(self):
        try:
            data = self.collect_settings()
            save_settings(data)
 
            result = update_report(
                export_path=data["export_path"],
                report_path=data["report_path"]
            )
 
            self.status_var.set(
                f"Отчет обновлен. Новых: {result['added_count']}, обновлено: {result['updated_count']}, удалено старых: {result['removed_count']}"
            )
 
            messagebox.showinfo(
                "Успешно",
                f"Отчет обновлен.\n"
                f"Строк из выгрузки после фильтра: {result['rows_in_export']}\n"
                f"Новых артикулов: {result['added_count']}\n"
                f"Обновленных артикулов: {result['updated_count']}\n"
                f"Удалено старых строк: {result['removed_count']}\n"
                f"Доп. колонок найдено: {result['extra_cols_count']}\n"
                f"Дата выгрузки: {result['export_modified_dt'].strftime('%d.%m.%Y %H:%M')}"
            )
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
 
def run_silent_update():
    try:
        s = load_settings()
        export_path = s.get("export_path", "").strip()
        report_path = s.get("report_path", "").strip()
 
        if not export_path or not report_path:
            return
        if not os.path.exists(export_path):
            return
        if not os.path.exists(report_path):
            return
 
        update_report(export_path, report_path)
    except Exception:
        pass
 
if __name__ == "__main__":
    if "--autoupdate" in sys.argv:
        run_silent_update()
    else:
        app = App()
        app.mainloop()